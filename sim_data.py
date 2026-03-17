import json
import os
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import openpyxl
from openpyxl.styles import Font
import urllib.request
import urllib.error
import subprocess
import sys

# Version information
VERSION = "1.0.0"
UPDATE_URL = "https://api.github.com/repos/YOUR_USERNAME/YOUR_REPO/releases/latest"  # Replace with your GitHub repo
UPDATE_CHECK_URL = "file://" + os.path.join(os.path.dirname(__file__), "version.json")  # Local file for testing

# Add new SIM

# Add new SIM
def add_sim():
    iccid = iccid_entry.get().strip()
    sim_class = type_var.get()
    carrier = carrier_var.get()
    if iccid=="" or iccid in existing:
        iccid_entry.delete(0, tk.END)
        iccid_entry.focus_set()
        return
    sim_type = detect_type(iccid, sim_class)
    sim = {
        "id": len(sims)+1, "iccid": iccid, "type": sim_type,
        "carrier": carrier, "status":"Korek", "phone_number":None,
        "date_added": datetime.now().isoformat(), "date_sold": None
    }
    sims.append(sim)
    existing.add(iccid)
    add_to_history(sim)
    iccid_entry.delete(0, tk.END)
    iccid_entry.focus_set()
    update_counts()
    save_data()

# Bulk add SIMs
def bulk_add():
    bulk_text = simpledialog.askstring("Bulk Add SIMs", "Paste ICCIDs, one per line:")
    if bulk_text:
        iccids = [line.strip() for line in bulk_text.split('\n') if line.strip()]
        added = 0
        for iccid in iccids:
            if iccid and iccid not in existing:
                sim_class = type_var.get()
                carrier = carrier_var.get()
                sim_type = detect_type(iccid, sim_class)
                sim = {
                    "id": len(sims)+1, "iccid": iccid, "type": sim_type,
                    "carrier": carrier, "status":"blank", "phone_number":None,
                    "date_added": datetime.now().isoformat(), "date_sold": None
                }
                sims.append(sim)
                existing.add(iccid)
                add_to_history(sim)
                added += 1
        if added > 0:
            update_counts()
            save_data()
            messagebox.showinfo("Bulk Add", f"Added {added} SIMs.")
        else:
            messagebox.showinfo("Bulk Add", "No new SIMs added (duplicates or empty).")

# Save file in the same directory as the program
FILE = os.path.join(os.path.dirname(__file__), "sim_inventory.json")

# Load data
def load_data():
    if not os.path.exists(FILE):
        return {"sims": [], "total_sold_ever": 0, "total_damaged_ever": 0}
    with open(FILE, "r") as f:
        loaded = json.load(f)
    if isinstance(loaded, list):
        # Migrate old data
        for sim in loaded:
            sim.setdefault("type", "U Kurdistan")
            sim.setdefault("carrier", "Korek")
            sim.setdefault("status", "blank")
            sim.setdefault("phone_number", None)
            sim.setdefault("date_added", datetime.now().isoformat())
            sim.setdefault("date_sold", None)
        return {"sims": loaded, "total_sold_ever": 0, "total_damaged_ever": 0}
    else:
        loaded.setdefault("total_sold_ever", 0)
        loaded.setdefault("total_damaged_ever", 0)
        for sim in loaded["sims"]:
            sim.setdefault("type", "U Kurdistan")
            sim.setdefault("carrier", "Korek")
            sim.setdefault("status", "blank")
            sim.setdefault("phone_number", None)
            sim.setdefault("date_added", datetime.now().isoformat())
            sim.setdefault("date_sold", None)
        return loaded

def save_data():
    with open(FILE, "w") as f:
        json.dump(data, f, indent=4)

data = load_data()
sims = data["sims"]
existing = {sim["iccid"] for sim in sims}

# Detect region from ICCID
def detect_type(iccid, sim_class):
    if iccid.startswith("892140800012"):
        region = "Kurdistan"
    elif iccid.startswith("892140800022"):
        region = "Baghdad"
    else:
        region = "Unknown"
    return f"{sim_class} {region}"

# Update counters safely
def update_counts():
    counts = {
        "U Kurdistan":0, "U Baghdad":0, "E Kurdistan":0, "E Baghdad":0,
        "U Unknown":0, "E Unknown":0, "Sold":0, "Damaged":0
    }
    for sim in sims:
        if sim["status"] == "sold":
            counts["Sold"] +=1
        elif sim["status"] == "damaged":
            counts["Damaged"] +=1
        else:
            counts.setdefault(sim["type"],0)
            counts[sim["type"]] +=1
    for key,lbl in count_labels.items():
        if key in counts:
            lbl.config(text=f"{key}: {counts[key]}")
        elif key == "Total Sold Ever":
            lbl.config(text=f"Total Sold Ever: {data['total_sold_ever']}")
        elif key == "Total Damaged Ever":
            lbl.config(text=f"Total Damaged Ever: {data['total_damaged_ever']}")

# Add to history table
def add_to_history(sim):
    history_tree.insert("", "end", values=(
        sim["id"], sim["iccid"], sim["type"], sim["carrier"], sim["status"],
        sim["phone_number"] or "", sim["date_added"].replace("T"," "),
        sim["date_sold"].replace("T"," ") if sim["date_sold"] else ""
    ))

# Edit selected SIM
def edit_sim():
    selected = history_tree.selection()
    if not selected:
        messagebox.showwarning("Edit SIM", "Please select a SIM to edit.")
        return
    iid = selected[0]
    sim_id = int(history_tree.item(iid, "values")[0])
    sim = next((s for s in sims if s["id"] == sim_id), None)
    if not sim:
        return

    # Create edit dialog
    edit_win = tk.Toplevel(root)
    edit_win.title("Edit SIM")
    edit_win.geometry("300x250")

    tk.Label(edit_win, text="ICCID:").grid(row=0, column=0, padx=5, pady=5)
    iccid_var = tk.StringVar(value=sim["iccid"])
    tk.Entry(edit_win, textvariable=iccid_var).grid(row=0, column=1, padx=5, pady=5)

    tk.Label(edit_win, text="Type:").grid(row=1, column=0, padx=5, pady=5)
    type_var_edit = tk.StringVar(value=sim["type"].split()[0] if " " in sim["type"] else sim["type"])
    tk.Entry(edit_win, textvariable=type_var_edit).grid(row=1, column=1, padx=5, pady=5)

    tk.Label(edit_win, text="Carrier:").grid(row=2, column=0, padx=5, pady=5)
    carrier_var_edit = tk.StringVar(value=sim["carrier"])
    tk.Entry(edit_win, textvariable=carrier_var_edit).grid(row=2, column=1, padx=5, pady=5)

    tk.Label(edit_win, text="Status:").grid(row=3, column=0, padx=5, pady=5)
    status_var = tk.StringVar(value=sim["status"])
    tk.Entry(edit_win, textvariable=status_var).grid(row=3, column=1, padx=5, pady=5)

    def save_edit():
        new_iccid = iccid_var.get().strip()
        new_type = type_var_edit.get().strip()
        new_carrier = carrier_var_edit.get().strip()
        new_status = status_var.get().strip()

        if new_iccid != sim["iccid"] and new_iccid in existing:
            messagebox.showerror("Error", "ICCID already exists.")
            return

        # Update existing set
        existing.discard(sim["iccid"])
        existing.add(new_iccid)

        # Update sim
        sim["iccid"] = new_iccid
        sim["type"] = detect_type(new_iccid, new_type)
        sim["carrier"] = new_carrier
        sim["status"] = new_status

        save_data()
        refresh_history()
        update_counts()
        edit_win.destroy()
        messagebox.showinfo("Edit SIM", "SIM updated successfully.")

    tk.Button(edit_win, text="Save", command=save_edit).grid(row=4, column=0, columnspan=2, pady=10)

# Mark sold with phone number
def sell_sim():
    selected = history_tree.selection()
    if not selected: return
    for iid in selected:
        sim_id = int(history_tree.item(iid,"values")[0])
        for sim in sims:
            if sim["id"]==sim_id:
                sim["status"]="sold"
                sim["date_sold"]=datetime.now().isoformat()
                data["total_sold_ever"] += 1
                phone = simpledialog.askstring("Assign Number",
                                               f"Assign phone number for ICCID {sim['iccid']}:")
                if phone:
                    sim["phone_number"]=phone
    save_data()
    refresh_history()
    update_counts()

# Mark damaged
def damage_sim():
    selected = history_tree.selection()
    if not selected: return
    for iid in selected:
        sim_id = int(history_tree.item(iid,"values")[0])
        for sim in sims:
            if sim["id"]==sim_id:
                sim["status"]="damaged"
                data["total_damaged_ever"] += 1
    save_data()
    refresh_history()
    update_counts()

# Delete selected SIM(s)
def delete_sim():
    selected = history_tree.selection()
    if not selected: return
    for iid in selected:
        sim_id = int(history_tree.item(iid,"values")[0])
        for i, sim in enumerate(sims):
            if sim["id"]==sim_id:
                sims.pop(i)
                existing.discard(sim["iccid"])
                break
    save_data()
    refresh_history()
    update_counts()

# Reset all history
def reset_history():
    global sims, existing
    if messagebox.askyesno("Confirm Reset","Are you sure you want to delete ALL SIM history?"):
        sims=[]
        existing=set()
        data["total_sold_ever"]=0
        data["total_damaged_ever"]=0
        save_data()
        refresh_history()
        update_counts()

# Refresh table
def refresh_history():
    for row in history_tree.get_children():
        history_tree.delete(row)
    for sim in sims:
        add_to_history(sim)

# Search SIMs
def search_sim():
    query = search_var.get().strip()
    for row in history_tree.get_children():
        history_tree.delete(row)
    for sim in sims:
        if query in sim["iccid"]:
            add_to_history(sim)

# Export unsold SIMs to Excel
def export_unsold():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unsold SIMs"
    headers = ["ID","ICCID","Type","Carrier","Status","Phone","Date Added","Date Sold"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for sim in sims:
        if sim["status"]!="sold":
            ws.append([
                sim["id"], sim["iccid"], sim["type"], sim["carrier"], sim["status"],
                sim["phone_number"] or "", sim["date_added"].replace("T"," "),
                sim["date_sold"].replace("T"," ") if sim["date_sold"] else ""
            ])
    save_path = os.path.join(os.path.dirname(__file__), "unsold_sims.xlsx")
    wb.save(save_path)
    messagebox.showinfo("Export Complete", f"Unsold SIMs exported to {save_path}")

# Import from Excel
def import_from_excel():
    from tkinter import filedialog
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            if len(row) < 2:
                continue
            iccid = str(row[1]).strip()  # Assuming ICCID in column 2
            if iccid and iccid not in existing:
                sim_type = detect_type(iccid, row[2] if len(row) > 2 and row[2] else "U")
                sim = {
                    "id": len(sims) + 1,
                    "iccid": iccid,
                    "type": sim_type,
                    "carrier": row[3] if len(row) > 3 and row[3] else "Korek",
                    "status": row[4] if len(row) > 4 and row[4] else "blank",
                    "phone_number": row[5] if len(row) > 5 and row[5] else None,
                    "date_added": row[6] if len(row) > 6 and row[6] else datetime.now().isoformat(),
                    "date_sold": row[7] if len(row) > 7 and row[7] else None
                }
                sims.append(sim)
                existing.add(iccid)
                add_to_history(sim)
                imported += 1
        if imported > 0:
            update_counts()
            save_data()
            messagebox.showinfo("Import Complete", f"Imported {imported} SIMs.")
        else:
            messagebox.showinfo("Import Complete", "No new SIMs imported (duplicates or invalid data).")
    except Exception as e:
        messagebox.showerror("Import Error", f"Error importing file: {str(e)}")

# Backup data
def backup_data():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = os.path.join(os.path.dirname(__file__), f"sim_inventory_backup_{timestamp}.json")
    import shutil
    shutil.copy(FILE, backup_file)
    messagebox.showinfo("Backup Complete", f"Backup saved as {backup_file}")

# Summary statistics
def show_summary():
    total = len(sims)
    sold = sum(1 for s in sims if s["status"] == "sold")
    damaged = sum(1 for s in sims if s["status"] == "damaged")
    blank = sum(1 for s in data if s["status"] == "blank")

    summary_win = tk.Toplevel(root)
    summary_win.title("Summary Statistics")
    summary_win.geometry("250x150")

    tk.Label(summary_win, text=f"Total SIMs: {total}", font=("Arial", 12)).pack(pady=5)
    tk.Label(summary_win, text=f"Sold: {sold}", font=("Arial", 12)).pack(pady=5)
    tk.Label(summary_win, text=f"Damaged: {damaged}", font=("Arial", 12)).pack(pady=5)
    tk.Label(summary_win, text=f"Blank: {blank}", font=("Arial", 12)).pack(pady=5)

# Update functions
def check_for_updates():
    """Check for available updates"""
    try:
        status_label.config(text="Checking for updates...")
        root.update()

        # Check local version file first
        version_file = os.path.join(os.path.dirname(__file__), "version.json")
        if os.path.exists(version_file):
            with open(version_file, 'r') as f:
                data = json.load(f)
        else:
            # Fallback to URL
            req = urllib.request.Request(UPDATE_CHECK_URL, headers={'User-Agent': 'SIM-Inventory-App'})
            with urllib.request.urlopen(req, timeout=10) as response:
                data = json.loads(response.read().decode())

        latest_version = data.get("version", VERSION)
        download_url = data.get("download_url", "")

        if latest_version > VERSION:
            changelog = data.get("changelog", [])
            changelog_text = "\n".join(f"• {item}" for item in changelog)

            if messagebox.askyesno("Update Available",
                                 f"A new version is available: {latest_version}\n"
                                 f"Current version: {VERSION}\n\n"
                                 f"What's new:\n{changelog_text}\n\n"
                                 "Would you like to download and install it?"):
                download_and_install_update(download_url, latest_version)
        else:
            messagebox.showinfo("Up to Date", f"You have the latest version: {VERSION}")

    except Exception as e:
        messagebox.showerror("Update Check Failed",
                           f"Could not check for updates.\nError: {str(e)}")
    finally:
        status_label.config(text="Ready")

def download_and_install_update(download_url, new_version):
    """Download and install the update"""
    try:
        status_label.config(text="Downloading update...")
        root.update()

        if download_url.startswith("file://"):
            # Local file for testing
            local_path = download_url.replace("file://", "")
            with open(local_path, 'r', encoding='utf-8') as f:
                new_code = f.read()
        else:
            # Download from URL
            req = urllib.request.Request(download_url, headers={'User-Agent': 'SIM-Inventory-App'})
            with urllib.request.urlopen(req, timeout=30) as response:
                new_code = response.read().decode()

        # Backup current file
        backup_file = FILE.replace('.json', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json')
        if os.path.exists(FILE):
            import shutil
            shutil.copy2(FILE, backup_file)

        # Create backup of current script
        script_backup = __file__ + f'.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        shutil.copy2(__file__, script_backup)

        # Write new version
        with open(__file__, 'w', encoding='utf-8') as f:
            f.write(new_code)

        messagebox.showinfo("Update Complete",
                          f"Update to version {new_version} installed successfully!\n\n"
                          "The application will now restart to apply the changes.")

        restart_application()

    except Exception as e:
        messagebox.showerror("Update Failed", f"Failed to install update: {str(e)}")
        status_label.config(text="Update failed")

def restart_application():
    """Restart the application"""
    try:
        # Save current data before restart
        save_data()

        # Start new instance
        if sys.platform == "win32":
            subprocess.Popen([sys.executable, __file__])
        else:
            subprocess.Popen([sys.executable, __file__])

        # Close current instance
        root.quit()
        root.destroy()
        sys.exit(0)

    except Exception as e:
        messagebox.showerror("Restart Failed", f"Could not restart application: {str(e)}")

def show_about():
    """Show about dialog"""
    about_text = f"""
SIM Inventory Management System
Version: {VERSION}

A comprehensive tool for managing SIM card inventory
with advanced tracking and reporting features.

Features:
• Add/Edit/Delete SIM cards
• Bulk operations
• Status tracking (Sold, Damaged, Available)
• Excel import/export
• Inline editing
• Keyboard shortcuts
• Auto-update functionality

© 2024 SIM Inventory Team
"""
    messagebox.showinfo("About", about_text.strip())

# GUI setup
root = tk.Tk()
root.title(f"SIM Inventory Management System v{VERSION}")
root.state('zoomed')  # Maximize window for any screen size
root.configure(bg="#ffffff")  # Clean white background

# Set ttk style for light theme
style = ttk.Style()
style.theme_use("clam")
style.configure("TButton", font=("Arial", 10), padding=5, background="#f8f9fa", foreground="#212529", borderwidth=1, relief="flat")
style.map("TButton", background=[("active", "#e9ecef")])
style.configure("Primary.TButton", background="#007bff", foreground="white")
style.map("Primary.TButton", background=[("active", "#0056b3")])
style.configure("Secondary.TButton", background="#6c757d", foreground="white")
style.map("Secondary.TButton", background=[("active", "#545b62")])
style.configure("TEntry", fieldbackground="#ffffff", borderwidth=1)
style.configure("TCombobox", fieldbackground="#ffffff", borderwidth=1)
style.configure("Treeview", background="#ffffff", fieldbackground="#ffffff", foreground="#212529")
style.map("Treeview", background=[("selected", "#007bff")], foreground=[("selected", "white")])

# Input frame
style = ttk.Style()
style.theme_use("clam")
style.configure("TButton", font=("Arial", 10), padding=5, background="#f8f9fa", foreground="#212529", borderwidth=1, relief="flat")
style.configure("Primary.TButton", background="#007bff", foreground="white", font=("Arial", 10, "bold"), relief="raised", borderwidth=2)
style.map("Primary.TButton", background=[("active", "#0056b3")])
style.configure("Secondary.TButton", background="#6c757d", foreground="white", font=("Arial", 10))
style.map("Secondary.TButton", background=[("active", "#5a6268")])
style.configure("TLabel", font=("Arial", 10), background="#ffffff", foreground="#212529")
style.configure("TEntry", font=("Arial", 12), fieldbackground="#ffffff", foreground="#212529", insertcolor="#007bff")
style.configure("Treeview", font=("Arial", 10), rowheight=25, background="#ffffff", foreground="#212529", fieldbackground="#ffffff")
style.configure("Treeview.Heading", font=("Arial", 11, "bold"), background="#007bff", foreground="white")
style.map("Treeview", background=[("selected", "#007bff")])
style.configure("TCombobox", fieldbackground="#ffffff", background="#ffffff", foreground="#212529")

# Top frame with title
frame_top = tk.Frame(root, bg="#007bff", padx=20, pady=10)
frame_top.pack(fill="x")
title_label = tk.Label(frame_top, text="SIM Inventory Management", font=("Arial", 18, "bold"), fg="white", bg="#007bff")
title_label.pack()

# Input frame
input_frame = tk.Frame(root, bg="#ffffff", padx=20, pady=10)
input_frame.pack(fill="x")

tk.Label(input_frame, text="ICCID:", font=("Arial", 12), bg="#ffffff", fg="#212529").grid(row=0, column=0, padx=10, pady=5, sticky="e")
iccid_entry = tk.Entry(input_frame, font=("Arial", 12), bg="#ffffff", fg="#212529", insertbackground="#007bff", width=25)
iccid_entry.grid(row=0, column=1, padx=10, pady=5)
iccid_entry.focus()
ttk.Button(input_frame, text="Bulk Add", command=bulk_add, style="Primary.TButton").grid(row=0, column=2, padx=10, pady=5)

tk.Label(input_frame, text="Type:", font=("Arial", 12), bg="#ffffff", fg="#212529").grid(row=0, column=3, padx=10, pady=5, sticky="e")
type_var = tk.StringVar()
type_menu = ttk.Combobox(input_frame, textvariable=type_var, values=["U", "E"], state="readonly", font=("Arial", 12))
type_menu.grid(row=0, column=4, padx=10, pady=5)
type_menu.current(0)

tk.Label(input_frame, text="Carrier:", font=("Arial", 12), bg="#ffffff", fg="#212529").grid(row=0, column=5, padx=10, pady=5, sticky="e")
carrier_var = tk.StringVar()
carrier_entry = tk.Entry(input_frame, font=("Arial", 12), bg="#ffffff", fg="#212529", insertbackground="#007bff", width=15)
carrier_entry.grid(row=0, column=6, padx=10, pady=5)
carrier_var.set("Korek")

iccid_entry.bind("<Return>", lambda e: add_sim())

# Search frame
search_frame = tk.Frame(root, bg="#ffffff", padx=20, pady=5)
search_frame.pack(fill="x")
tk.Label(search_frame, text="Search ICCID:", font=("Arial", 12), bg="#ffffff", fg="#212529").grid(row=0, column=0, padx=10, pady=5, sticky="e")
search_var = tk.StringVar()
search_entry = tk.Entry(search_frame, textvariable=search_var, font=("Arial", 12), bg="#ffffff", fg="#212529", insertbackground="#007bff", width=25)
search_entry.grid(row=0, column=1, padx=10, pady=5)
ttk.Button(search_frame, text="Search", command=search_sim, style="Secondary.TButton").grid(row=0, column=2, padx=10, pady=5)

# Counts
frame_counts = tk.Frame(root, bg="#ffffff", padx=20, pady=5)
frame_counts.pack(fill="x")
count_labels = {}
for i,key in enumerate(["U Kurdistan","U Baghdad","E Kurdistan","E Baghdad","U Unknown","E Unknown","Sold","Damaged","Total Sold Ever","Total Damaged Ever"]):
    lbl = tk.Label(frame_counts,text=f"{key}: 0",font=("Arial",12), bg="#ffffff", fg="#212529")
    lbl.grid(row=0,column=i,padx=10)
    count_labels[key]=lbl

# History table
table_frame = tk.Frame(root, bg="#ffffff", padx=20, pady=10)
table_frame.pack(fill="both", expand=True)
columns = ("ID","ICCID","Type","Carrier","Status","Phone","Date Added","Date Sold")
history_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=20)
for col in columns:
    history_tree.heading(col, text=col)
    history_tree.column(col, width=120, stretch=True)  # Allow columns to stretch
history_tree.pack(fill="both", expand=True, side="left")

# Scrollbar for treeview
scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=history_tree.yview)
history_tree.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side="right", fill="y")

# Inline editing variables
edit_entry = None
edit_item = None
edit_column = None

def start_inline_edit(event):
    global edit_entry, edit_item, edit_column
    # Get the item and column clicked
    region = history_tree.identify_region(event.x, event.y)
    if region != "cell":
        return

    column = history_tree.identify_column(event.x)
    item = history_tree.identify_row(event.y)

    if not item or column == "#0":  # Skip if no item or tree column
        return

    # Get column index
    col_index = int(column[1:]) - 1  # #1 -> 0, #2 -> 1, etc.
    if col_index >= len(columns):
        return

    # Don't allow editing ID, Date Added, Date Sold
    if columns[col_index] in ["ID", "Date Added", "Date Sold"]:
        return

    edit_item = item
    edit_column = col_index

    # Get current value
    current_value = history_tree.item(item, "values")[col_index]

    # Get cell bbox
    x, y, width, height = history_tree.bbox(item, column)

    # Create entry widget
    edit_entry = tk.Entry(table_frame, font=("Arial", 10))
    edit_entry.insert(0, current_value)
    edit_entry.select_range(0, tk.END)
    edit_entry.focus()

    # Position the entry
    edit_entry.place(x=x, y=y, width=width, height=height)

    # Bind events
    edit_entry.bind("<Return>", save_inline_edit)
    edit_entry.bind("<Escape>", cancel_inline_edit)
    edit_entry.bind("<FocusOut>", save_inline_edit)

def save_inline_edit(event=None):
    global edit_entry, edit_item, edit_column
    if not edit_entry or not edit_item:
        return

    new_value = edit_entry.get().strip()
    col_name = columns[edit_column]

    # Find the SIM
    sim_id = int(history_tree.item(edit_item, "values")[0])
    sim = next((s for s in sims if s["id"] == sim_id), None)
    if not sim:
        cancel_inline_edit()
        return

    # Update based on column
    if col_name == "ICCID":
        if new_value and new_value != sim["iccid"]:
            if new_value in existing:
                messagebox.showerror("Error", "ICCID already exists!")
                cancel_inline_edit()
                return
            existing.discard(sim["iccid"])
            existing.add(new_value)
            sim["iccid"] = new_value
    elif col_name == "Type":
        sim["type"] = detect_type(sim["iccid"], new_value)
    elif col_name == "Carrier":
        sim["carrier"] = new_value
    elif col_name == "Status":
        old_status = sim["status"]
        sim["status"] = new_value.lower()
        # Handle counter updates for status changes
        if old_status != "sold" and new_value.lower() == "sold":
            data["total_sold_ever"] += 1
        elif old_status != "damaged" and new_value.lower() == "damaged":
            data["total_damaged_ever"] += 1
    elif col_name == "Phone":
        sim["phone_number"] = new_value if new_value else None

    save_data()
    refresh_history()
    update_counts()
    cancel_inline_edit()

def cancel_inline_edit(event=None):
    global edit_entry
    if edit_entry:
        edit_entry.destroy()
        edit_entry = None

# Bind double-click for inline editing
history_tree.bind("<Double-1>", start_inline_edit)

def paste_to_entry():
    try:
        content = root.clipboard_get()
        iccid_entry.insert(tk.END, content)
    except tk.TclError:
        pass  # Clipboard empty, do nothing
def copy_iccid():
    selected = history_tree.selection()
    if selected:
        iid = selected[0]
        values = history_tree.item(iid, "values")
        root.clipboard_clear()
        root.clipboard_append(values[1])  # ICCID is index 1

def copy_row():
    selected = history_tree.selection()
    if selected:
        iid = selected[0]
        values = history_tree.item(iid, "values")
        text = "\t".join(str(v) for v in values)
        root.clipboard_clear()
        root.clipboard_append(text)

# Context menu
popup_menu = tk.Menu(root, tearoff=0)
popup_menu.add_command(label="Copy ICCID", command=copy_iccid)
popup_menu.add_command(label="Copy Row", command=copy_row)
popup_menu.add_separator()
popup_menu.add_command(label="Quick Edit", command=lambda: edit_sim())
popup_menu.add_command(label="Mark as Sold", command=lambda: quick_sell())
popup_menu.add_command(label="Mark as Damaged", command=lambda: quick_damage())
popup_menu.add_command(label="Assign Phone Number", command=lambda: quick_assign_phone())
popup_menu.add_separator()
popup_menu.add_command(label="Delete Selected", command=lambda: delete_sim())

def quick_sell():
    selected = history_tree.selection()
    if not selected: return
    for iid in selected:
        sim_id = int(history_tree.item(iid,"values")[0])
        for sim in sims:
            if sim["id"]==sim_id and sim["status"] != "sold":
                sim["status"]="sold"
                sim["date_sold"]=datetime.now().isoformat()
                data["total_sold_ever"] += 1
                phone = simpledialog.askstring("Assign Number",
                                               f"Assign phone number for ICCID {sim['iccid']}:")
                if phone:
                    sim["phone_number"]=phone
    save_data()
    refresh_history()
    update_counts()

def quick_damage():
    selected = history_tree.selection()
    if not selected: return
    for iid in selected:
        sim_id = int(history_tree.item(iid,"values")[0])
        for sim in sims:
            if sim["id"]==sim_id and sim["status"] != "damaged":
                sim["status"]="damaged"
                data["total_damaged_ever"] += 1
    save_data()
    refresh_history()
    update_counts()

def quick_assign_phone():
    selected = history_tree.selection()
    if not selected:
        messagebox.showwarning("Assign Phone", "Please select SIMs to assign phone numbers.")
        return
    phone = simpledialog.askstring("Assign Phone Number", "Enter phone number:")
    if phone:
        for iid in selected:
            sim_id = int(history_tree.item(iid,"values")[0])
            for sim in sims:
                if sim["id"]==sim_id:
                    sim["phone_number"]=phone
        save_data()
        refresh_history()
        update_counts()

def show_popup(e):
    popup_menu.post(e.x_root, e.y_root)

# Bind right-click to show menu
history_tree.bind("<Button-3>", show_popup)
# Bind Ctrl+C to copy row
history_tree.bind("<Control-c>", lambda e: copy_row())
# Keyboard shortcuts for quick updates
root.bind("<Control-s>", lambda e: quick_sell())
root.bind("<Control-d>", lambda e: quick_damage())
root.bind("<Control-e>", lambda e: edit_sim())
root.bind("<Delete>", lambda e: delete_sim())

# Bottom buttons
button_frame = tk.Frame(root, bg="#ffffff", padx=20, pady=10)
button_frame.pack(fill="x")
buttons = [
    ("Edit Selected", edit_sim, "Primary.TButton"),
    ("Mark Sold", sell_sim, "Primary.TButton"),
    ("Mark Damaged", damage_sim, "Primary.TButton"),
    ("Delete Selected", delete_sim, "Primary.TButton"),
    ("Backup Data", backup_data, "Secondary.TButton"),
    ("Summary", show_summary, "Secondary.TButton"),
    ("Reset All", reset_history, "Secondary.TButton"),
    ("Import from Excel", import_from_excel, "Secondary.TButton"),
    ("Export Unsold to Excel", export_unsold, "Secondary.TButton")
]
for i, (text, cmd, btn_style) in enumerate(buttons):
    ttk.Button(button_frame, text=text, command=cmd, style=btn_style).grid(row=0, column=i, padx=5, pady=5)

# Status bar
status_frame = tk.Frame(root, bg="#f8f9fa", padx=20, pady=5)
status_frame.pack(fill="x")
path_label = tk.Label(status_frame, text=f"Saving to: {FILE}", font=("Arial", 9), bg="#f8f9fa", fg="#6c757d")
path_label.pack(side="left")
status_label = tk.Label(status_frame, text="Ready", font=("Arial", 9), bg="#f8f9fa", fg="#6c757d")
status_label.pack(side="right")

def show_shortcuts():
    shortcuts = """
Keyboard Shortcuts:
• Double-click cell: Inline edit
• Right-click: Context menu
• Ctrl+S: Mark selected as Sold
• Ctrl+D: Mark selected as Damaged
• Ctrl+E: Edit selected SIM
• Delete: Delete selected SIMs
• Ctrl+C: Copy selected row

Inline Editing:
• Double-click any cell (except ID/Date fields)
• Press Enter or click outside to save
• Press Escape to cancel
"""
    messagebox.showinfo("Keyboard Shortcuts & Tips", shortcuts.strip())

help_button = ttk.Button(status_frame, text="?", width=3, command=show_shortcuts, style="Secondary.TButton")
help_button.pack(side="right", padx=(0, 10))

# Initialize
refresh_history()
update_counts()
root.mainloop()